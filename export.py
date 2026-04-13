"""
export.py — Export des résultats d'analyse en Excel (.xlsx) ou CSV.

Génère un classeur Excel avec plusieurs onglets :
1. Synthèse       — récapitulatif multi-tickers (signal, scores, indicateurs clés)
2. Technique      — scorecard détaillée par ticker
3. Évolution 3M   — performance glissante, volatilité, drawdown
4. Fondamentaux   — critères et score fondamental
5. OHLCV          — données brutes par ticker (derniers 90j)
6. Actualités     — news agrégées par ticker
"""

import io
import logging
from datetime import datetime
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)


# ─── Couleurs pour la mise en forme Excel ─────────────────────────────────────

COLOR_ACHAT  = "C8F0E3"  # Vert clair
COLOR_VENTE  = "FADADA"  # Rouge clair
COLOR_NEUTRE = "FFF3CD"  # Jaune clair
COLOR_HEADER = "2D4A3E"  # Vert foncé BRVM
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_SOLIDE = "C8F0E3"
COLOR_FAIBLE = "FADADA"


def generate_excel(
    results: dict,
    include_ohlcv: bool = True,
    include_news: bool = True,
) -> bytes:
    """
    Génère un fichier Excel en mémoire et retourne les bytes.

    Args:
        results:       Dict {ticker: result_dict} issu du pipeline analyser_ticker
        include_ohlcv: Inclure l'onglet OHLCV
        include_news:  Inclure l'onglet Actualités

    Returns:
        Bytes du fichier .xlsx prêt à être téléchargé
    """
    output = io.BytesIO()

    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_synthese(writer, results)
            _write_scorecard_technique(writer, results)
            _write_evolution(writer, results)
            _write_fondamentaux(writer, results)
            if include_ohlcv:
                _write_ohlcv(writer, results)
            if include_news:
                _write_news(writer, results)

            _style_workbook(writer)

    except Exception as e:
        logger.exception(f"[Export] Erreur génération Excel: {e}")
        raise

    return output.getvalue()


def generate_csv(results: dict) -> bytes:
    """
    Génère un CSV de synthèse simple (pour téléchargement rapide).
    Retourne les bytes encodés en UTF-8 avec BOM pour compatibilité Excel FR.
    """
    rows = _build_synthese_rows(results)
    df = pd.DataFrame(rows)
    return df.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")


# ─── Onglet 1 : Synthèse ─────────────────────────────────────────────────────

def _write_synthese(writer: pd.ExcelWriter, results: dict) -> None:
    rows = _build_synthese_rows(results)
    if not rows:
        return
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Synthèse", index=False)
    _add_metadata_row(writer.sheets["Synthèse"])


def _build_synthese_rows(results: dict) -> list[dict]:
    rows = []
    for ticker, result in results.items():
        if result is None:
            continue
        ind = result.get("ind")
        score = result.get("score")
        fd = result.get("fundamentals")
        evo = result.get("evolution")

        if ind is None or score is None:
            continue

        row = {
            "Ticker": ticker,
            "Cours (FCFA)": f"{ind.cours_actuel:,.0f}" if ind.cours_actuel else "N/D",
            "Var J-1 (%)": f"{ind.variation_j1_pct:+.2f}%" if ind.variation_j1_pct is not None else "N/D",
            "Signal Technique": f"{score.signal_emoji} {score.signal}",
            "Score Technique": f"{score.score_total:+d}",
            "Confiance": score.confiance.capitalize(),
            "RSI": f"{ind.rsi:.1f}" if ind.rsi else "N/D",
            "MACD": ind.macd_signal.capitalize() if ind.macd_signal else "N/D",
            "ADX": f"{ind.adx:.0f}" if ind.adx else "N/D",
            "Perf 1M (%)": f"{ind.perf_1m:+.1f}%" if ind.perf_1m is not None else "N/D",
            "Perf 3M (%)": f"{ind.perf_3m:+.1f}%" if ind.perf_3m is not None else "N/D",
            "Alpha 1M vs BRVMC": f"{ind.perf_vs_index_1m:+.1f}%" if ind.perf_vs_index_1m is not None else "N/D",
        }

        if evo is not None:
            row["Volatilité (%)"] = f"{evo.volatilite_63j:.1f}%" if evo.volatilite_63j else "N/D"
            row["Max Drawdown (%)"] = f"{evo.max_drawdown_pct:.1f}%" if evo.max_drawdown_pct else "N/D"

        if fd is not None and fd.donnees_disponibles:
            row["Signal Fondamental"] = f"{fd.signal_emoji} {fd.signal_fondamental}"
            row["Score Fondamental"] = f"{fd.score_fondamental:+d}"
            row["PER"] = f"{fd.per:.1f}×" if fd.per else "N/D"
            row["Rdt Dividende (%)"] = f"{fd.rendement_dividende_pct:.1f}%" if fd.rendement_dividende_pct else "N/D"
            if fd.capitalisation_fcfa:
                cap_str = _format_large_number(fd.capitalisation_fcfa)
                row["Capitalisation"] = cap_str

        rows.append(row)
    return rows


# ─── Onglet 2 : Scorecard technique ──────────────────────────────────────────

def _write_scorecard_technique(writer: pd.ExcelWriter, results: dict) -> None:
    rows = []
    for ticker, result in results.items():
        if result is None:
            continue
        score = result.get("score")
        if score is None:
            continue
        for critere in score.criteres:
            rows.append({
                "Ticker": ticker,
                "Critère": critere.nom,
                "Valeur": critere.valeur,
                "Points": f"{critere.points:+d}",
                "Interprétation": critere.interpretation,
            })
        rows.append({
            "Ticker": ticker,
            "Critère": "─── TOTAL ───",
            "Valeur": "",
            "Points": f"{score.score_total:+d}",
            "Interprétation": f"Signal : {score.signal} | Confiance : {score.confiance}",
        })

    if not rows:
        return
    pd.DataFrame(rows).to_excel(writer, sheet_name="Technique", index=False)


# ─── Onglet 3 : Évolution 3M ─────────────────────────────────────────────────

def _write_evolution(writer: pd.ExcelWriter, results: dict) -> None:
    rows_perf = []
    rows_events = []

    for ticker, result in results.items():
        if result is None:
            continue
        evo = result.get("evolution")
        if evo is None:
            continue

        # Performances glissantes
        for p in evo.periodes_glissantes:
            rows_perf.append({
                "Ticker": ticker,
                "Période": p.label,
                "Début": p.debut.strftime("%d/%m/%Y"),
                "Fin": p.fin.strftime("%d/%m/%Y"),
                "Performance (%)": f"{p.perf_pct:+.2f}%",
                "Alpha vs BRVMC (%)": f"{p.vs_index_pct:+.2f}%" if p.vs_index_pct is not None else "N/D",
            })

        # Résumé volatilité / drawdown
        rows_perf.append({
            "Ticker": ticker,
            "Période": "─── STATS ───",
            "Début": "",
            "Fin": "",
            "Performance (%)": f"Volatilité 3M : {evo.volatilite_63j:.1f}%" if evo.volatilite_63j else "N/D",
            "Alpha vs BRVMC (%)": f"Max DD : {evo.max_drawdown_pct:.1f}%" if evo.max_drawdown_pct else "N/D",
        })

        # Événements
        for ev in evo.evenements:
            rows_events.append({
                "Ticker": ticker,
                "Date": ev.date.strftime("%d/%m/%Y"),
                "Type": ev.type.replace("_", " ").title(),
                "Prix (FCFA)": f"{ev.prix:,.0f}" if ev.prix else "N/D",
                "Intensité": ev.intensite.capitalize(),
                "Description": ev.description,
            })

    if rows_perf:
        pd.DataFrame(rows_perf).to_excel(writer, sheet_name="Évolution 3M", index=False)
    if rows_events:
        pd.DataFrame(rows_events).to_excel(writer, sheet_name="Événements", index=False)


# ─── Onglet 4 : Fondamentaux ─────────────────────────────────────────────────

def _write_fondamentaux(writer: pd.ExcelWriter, results: dict) -> None:
    rows_summary = []
    rows_criteres = []

    for ticker, result in results.items():
        if result is None:
            continue
        fd = result.get("fundamentals")
        if fd is None or not fd.donnees_disponibles:
            continue

        rows_summary.append({
            "Ticker": ticker,
            "Signal": f"{fd.signal_emoji} {fd.signal_fondamental}",
            "Score": f"{fd.score_fondamental:+d}",
            "Capitalisation": _format_large_number(fd.capitalisation_fcfa) if fd.capitalisation_fcfa else "N/D",
            "PER": f"{fd.per:.1f}×" if fd.per else "N/D",
            "Dividende (FCFA)": f"{fd.dividende_par_action:,.0f}" if fd.dividende_par_action else "N/D",
            "Rdt Dividende (%)": f"{fd.rendement_dividende_pct:.1f}%" if fd.rendement_dividende_pct else "N/D",
            "Plus haut 52S": f"{fd.plus_haut_52s:,.0f}" if fd.plus_haut_52s else "N/D",
            "Plus bas 52S": f"{fd.plus_bas_52s:,.0f}" if fd.plus_bas_52s else "N/D",
            "Position 52S (%)": f"{fd.position_52s_pct:.0f}%" if fd.position_52s_pct is not None else "N/D",
            "Nb titres": _format_large_number(fd.nombre_titres) if fd.nombre_titres else "N/D",
            "Source": fd.source,
        })

        for critere in fd.criteres_fondamentaux:
            rows_criteres.append({
                "Ticker": ticker,
                "Critère": critere["nom"],
                "Valeur": critere["valeur"],
                "Points": f"{critere['points']:+d}",
                "Interprétation": critere["interpretation"],
            })

    if rows_summary:
        pd.DataFrame(rows_summary).to_excel(writer, sheet_name="Fondamentaux", index=False)
    if rows_criteres:
        pd.DataFrame(rows_criteres).to_excel(writer, sheet_name="Score Fondamental", index=False)


# ─── Onglet 5 : OHLCV ────────────────────────────────────────────────────────

def _write_ohlcv(writer: pd.ExcelWriter, results: dict) -> None:
    """Écrit les données OHLCV brutes des 90 derniers jours pour chaque ticker."""
    frames = []
    for ticker, result in results.items():
        if result is None:
            continue
        df = result.get("df")
        if df is None or df.empty:
            continue
        df_export = df.tail(90).copy()
        df_export.index = df_export.index.strftime("%d/%m/%Y")
        df_export.index.name = "Date"
        df_export.insert(0, "Ticker", ticker)
        frames.append(df_export.reset_index())

    if frames:
        pd.concat(frames).to_excel(writer, sheet_name="OHLCV", index=False)


# ─── Onglet 6 : Actualités ───────────────────────────────────────────────────

def _write_news(writer: pd.ExcelWriter, results: dict) -> None:
    rows = []
    for ticker, result in results.items():
        if result is None:
            continue
        for article in result.get("news", []):
            rows.append({
                "Ticker": ticker,
                "Date": article.get("date", ""),
                "Titre": article.get("titre", ""),
                "Source": article.get("source", ""),
                "Résumé": article.get("resume", ""),
                "URL": article.get("url", ""),
            })
    if rows:
        pd.DataFrame(rows).to_excel(writer, sheet_name="Actualités", index=False)


# ─── Mise en forme Excel ──────────────────────────────────────────────────────

def _style_workbook(writer: pd.ExcelWriter) -> None:
    """Applique une mise en forme basique aux onglets du classeur."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = writer.book
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_font = Font(color=COLOR_HEADER_TEXT, bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for ws in wb.worksheets:
            # En-tête
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            # Coloration conditionnelle colonne "Signal"
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "ACHAT" in cell.value or "SOLIDE" in cell.value:
                            cell.fill = PatternFill("solid", fgColor=COLOR_ACHAT)
                        elif "VENTE" in cell.value or "FAIBLE" in cell.value:
                            cell.fill = PatternFill("solid", fgColor=COLOR_VENTE)
                        elif "NEUTRE" in cell.value:
                            cell.fill = PatternFill("solid", fgColor=COLOR_NEUTRE)

            # Largeur des colonnes auto
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = max(
                    (len(str(c.value)) if c.value else 0 for c in col),
                    default=10
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

            # Figer la première ligne
            ws.freeze_panes = "A2"

    except ImportError:
        logger.debug("[Export] openpyxl non disponible pour le styling")
    except Exception as e:
        logger.debug(f"[Export] Erreur styling: {e}")


def _add_metadata_row(ws) -> None:
    """Ajoute une ligne de métadonnées en bas du premier onglet."""
    try:
        ws.append([])
        ws.append([f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — BRVM Screener — Investment Pioneers"])
        ws.append(["Données à titre informatif uniquement. Pas de conseil en investissement."])
    except Exception:
        pass


# ─── Utilitaires ─────────────────────────────────────────────────────────────

def _format_large_number(value: Optional[float]) -> str:
    """Formate un grand nombre en chaîne lisible (Mds, M, K)."""
    if value is None:
        return "N/D"
    if value >= 1e9:
        return f"{value / 1e9:.1f} Mds FCFA"
    if value >= 1e6:
        return f"{value / 1e6:.0f} M FCFA"
    if value >= 1e3:
        return f"{value / 1e3:.0f} K FCFA"
    return f"{value:,.0f} FCFA"
